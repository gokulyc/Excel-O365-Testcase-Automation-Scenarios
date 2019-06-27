from openpyxl.styles import NamedStyle, Font, Border, Side
import json
import xlsxwriter
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from pathlib import Path
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

with open('options.json') as f:
    data = json.load(f)

print(data['name'] + "  " + data['languages'][0])
wb = load_workbook(filename=Path(data['path1']))
#wb = Workbook('in\a.xlsx')
sheet_ranges = wb['Sheet1']
print(sheet_ranges['D18'].value)
ws = wb.active


#highlight1 = NamedStyle(name="highlight")
    #highlight1.font = Font(bold=True, size=12)
bd = Side(style='thick', color="000000")
    #highlight1.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    # wb.add
    # wb.add_named_style(highlight1)
#ws['C28'].border = Border(left=bd, top=bd, right=bd, bottom=bd)
#ws['E9'].font = Font(bold=True, size=12)
ws['C28']="=COUNTA(C8:C24)"
ws['C28'].style='60 % - Accent6'
i_val=ws['c28'].value
#copy values to cells E28, G28, I28, K28 & M28
ws['E28']=i_val
ws['G28']=i_val
ws['I28']=i_val
ws['K28']=i_val
ws['M28']=i_val
ws['E28'].alignment =Alignment(horizontal="center")
ws['G28'].alignment=Alignment(horizontal="center")
ws['I28'].alignment=Alignment(horizontal="center")
ws['K28'].alignment=Alignment(horizontal="center")
ws['M28'].alignment=Alignment(horizontal="center")
ws['E28'].style='60 % - Accent6'
ws['G28'].style='60 % - Accent6'
ws['I28'].style='60 % - Accent6'
ws['K28'].style='60 % - Accent6'
ws['M28'].style='60 % - Accent6'
ws['E28'].border=Border(outline=True)
ws['G28'].border=Border(outline=True)
ws['I28'].border=Border(outline=True)
ws['K28'].border=Border(outline=True)
ws['M28'].border=Border(outline=True)
ws['M28'].fill=PatternFill(bgColor='000000',fgColor='ffffff')
#ws['E9'].Side(style='thick', color="000000")



wb.save(filename=Path(data['path1']))
wb.close()
print("done and closing")

# Output: {'name': 'Bob', 'languages': ['English', 'Fench']}
# {
#    "name": "Bob",
#    "languages": [
#        "English",
#        "Fench"
#    ]
# }
